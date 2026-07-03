from pathlib import Path
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

FONT_PATHS = [
    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
    'C:/Windows/Fonts/msjh.ttc',
]


def text_width(draw, text, font):
    return int(draw.textbbox((0, 0), str(text), font=font)[2])


def format_value(v):
    if v is None:
        return ''
    if isinstance(v, float):
        return f'{v:.1f}'.rstrip('0').rstrip('.')
    # GitHub runners may not have emoji-capable fonts; use a plain marker.
    return str(v).replace('✅', '有')


def export_stock_image(xlsx_path, output_path, mode='limit'):
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = 'B_XGB獨立' if mode == 'xgb' else '漲停候選'
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'找不到「{sheet_name}」工作表')
    ws = wb[sheet_name]
    # 用表頭名稱找欄位，避免新增欄位時位置跑掉
    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def num(row, col):
        try:
            return float(ws.cell(row, col).value)
        except (TypeError, ValueError):
            return 0.0

    if mode == 'xgb':
        title = 'B XGB獨立'
        columns = [header[n] for n in ('等級', '代號', '股名', '有期貨', 'XGB信心%')]
        max_rows = 20
        row_filter = lambda r: True
    else:
        title = '核心漲停候選'
        columns = [header[n] for n in ('等級', '代號', '股名', '有期貨', 'XGB信心%', '漲停候選分數%')]
        max_rows = 8

        def row_filter(r):
            return (ws.cell(r, header['等級']).value == 'A 強訊號'
                    and num(r, header['XGB信心%']) >= 80
                    and num(r, header['漲停候選分數%']) >= 90)

    rows = [
        r for r in range(2, ws.max_row + 1)
        if row_filter(r)
    ][:max_rows]

    data = []
    for r in [1] + rows:
        data.append([format_value(ws.cell(r, c).value) for c in columns])
    if not rows:
        data.append(['無資料'] + [''] * (len(columns) - 1))

    font_path = next((p for p in FONT_PATHS if Path(p).exists()), None)
    if not font_path:
        raise FileNotFoundError('找不到中文字型')
    font = ImageFont.truetype(font_path, 22)
    header_font = ImageFont.truetype(font_path, 22)
    tmp_img = Image.new('RGB', (1, 1), 'white')
    draw = ImageDraw.Draw(tmp_img)

    # Auto-fit columns with padding. Keep it compact like the Excel crop.
    col_widths = []
    for c in range(len(columns)):
        widest = max(text_width(draw, row[c], header_font if i == 0 else font) for i, row in enumerate(data))
        col_widths.append(max(92, widest + 28))

    row_h = 34
    margin = 12
    title_h = 42
    w = sum(col_widths) + margin * 2
    h = title_h + row_h * len(data) + margin * 2

    img = Image.new('RGB', (w, h), 'white')
    draw = ImageDraw.Draw(img)
    draw.text(
        (margin, margin - 2),
        f'{Path(xlsx_path).stem}｜{title}',
        fill=(40, 40, 40),
        font=header_font
    )

    y0 = margin + title_h
    grid = (205, 205, 205)
    header_bg = (238, 242, 247)

    for r, row in enumerate(data):
        y = y0 + r * row_h
        x = margin
        for c, val in enumerate(row):
            if r == 0:
                draw.rectangle([x, y, x + col_widths[c], y + row_h], fill=header_bg)
            draw.rectangle([x, y, x + col_widths[c], y + row_h], outline=grid)
            draw.text((x + 8, y + 4), val, fill=(20, 20, 20), font=header_font if r == 0 else font)
            x += col_widths[c]

    img.save(output_path)
    return output_path


if __name__ == '__main__':
    import sys
    export_stock_image(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else 'limit')
